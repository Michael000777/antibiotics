#!/usr/bin/env python

import numpy as np
import argparse
import sys,math
import itertools
import openpyxl
import cairo

from skimage import measure


class PlateReaderData(object):
    def __init__(self,**kwargs):

        self.__infilenames = kwargs.get("infiles",[])
        if self.__infilenames is None:
            raise IOError('No input data provided. Use option -i FILENAME1 [FILENAME2 ...]')
        self.__ignoresheetsparameter = kwargs.get("ignoresheets",[])

        self.extract_figure_file_parameters(kwargs)

        self.__designdata = list()
        self.__designtitle = list()

        self.__data = list()
        self.__filenames = list()
        self.__sheetnames = list()
        self.__designassignment = list()
        
        self.__rescale = kwargs.get("DataRescale",False)
        self.__logscale = kwargs.get("DataLogscale",False)
        self.__logmin  = kwargs.get("DataLogscaleMin",-20)
        
        self.__read_coordinates = { 'x0':kwargs.get("xstart",2),
                                    'y0':kwargs.get("ystart",2),
                                    'xwidth':kwargs.get("xwidth",12),
                                    'yheight':kwargs.get("yheight",8) }
        
        self.__ignoresheets = ['Plate Design*']
        if len(self.__ignoresheetsparameter) > 0:
            self.__ignoresheets += self.__ignoresheetsparameter
        
        # load all data at __init__()
        if len(self.__infilenames) > 0:
            for fn in self.__infilenames:
                try:
                    data = openpyxl.load_workbook(fn)
                except:
                    continue
                
                for designsheet in [s for s in data if self.ignoresheet(s.title)]:
                    self.__designdata.append(self.read_initial_conditions(data,designsheet.title))
                    self.__designtitle.append(designsheet.title)
                i = 0
                for sheet in [s for s in data if not self.ignoresheet(s.title)]:
                    self.__data.append(self.read_sheet(sheet))
                    self.__sheetnames.append(sheet.title)
                    self.__filenames.append(fn)
                    try:
                        design = kwargs.get("designassignment",[])[i]
                        if not 0 <= design < len(self.__designdata):
                            raise KeyError
                        self.__designassignment.append(design)
                    except:
                        self.__designassignment.append(0)
                    i+=1
        else:
            raise IOError('No input data provided. Use option -i FILENAME1 [FILENAME2 ...]')
                
            
        
    
    def column_string(self,n):
        div=n
        string=""
        temp=0
        while div>0:
            module=(div-1)%26
            string=chr(65+module)+string
            div=int((div-module)/26)
        return string
    
    
    def rgb(self,color):
        r = int(color[0:2],16)
        g = int(color[2:4],16)
        b = int(color[4:6],16)
        return np.array([r/255.,g/255.,b/255.])


    def avg(self,values,geom = True):
        if len(values) > 0:
            if geom:    return np.power(np.product(values),1./len(values))
            else:       return np.mean(values)
        else:
            raise ValueError
    
    
    def read_sheet(self,sheet,x0 = None, y0 = None, width = None, height = None):
        if x0 is None:      x0 = self.__read_coordinates['x0']
        if y0 is None:      y0 = self.__read_coordinates['y0']
        if width is None:   width = self.__read_coordinates['xwidth']
        if height is None:  height = self.__read_coordinates['yheight']
        
        return np.array([[sheet['{:s}{:d}'.format(self.column_string(i),j)].value for i in range(x0,x0+width)] for j in range(y0,y0+height)],dtype=np.float)


    def read_initial_conditions(self,data,sheetname,xab = 4, yab = 14, xcells = 4, ycells = 3, width = 12, height = 8):
        if sheetname in data.sheetnames:
            return self.read_sheet(data[sheetname],x0 = xab,y0 = yab),self.read_sheet(data[sheetname],x0 = xcells,y0 = ycells)
        else:
            raise KeyError
    

    def ignoresheet(self,sheetname):
        r = False
        for ignore in self.__ignoresheets:
            if len(ignore) > 0:
                if ignore[-1] == '*':
                    if sheetname[:min(len(sheetname),len(ignore)-1)].upper() == ignore[:-1].upper():
                        r = True
                else:
                    if sheetname.upper() == ignore.upper():
                        r = True
            else:
                if sheetname.upper() == ignore.upper():
                    r = True
        return r
    
    
    def rescale(self,g):
        if self.__rescale:
            if self.__logscale:
                r = np.log(g)
                r[r<self.__logmin] = self.__logmin
            else:
                r = g[:,:]
                
            r = (r - np.min(r))/(np.max(r) - np.min(r))
            return r
        else:
            return g
    
    
    def extract_figure_file_parameters(self,kwargs):
        # default values
        self.figureparameters = {   'colors':   ['3465a4','ffffff','2e3436','eeeeec','a40000'],
                                    'wellradius': 20,
                                    'wellsize':50,
                                    'linewidth':3}
        # update default values if part of the argument dictionary
        if kwargs.has_key('FigureWellDistance'):        self.figureparameters['wellsize']   = kwargs['FigureWellDistance']
        if kwargs.has_key('FigureWellRadius'):          self.figureparameters['wellradius'] = kwargs['FigureWellRadius']
        if kwargs.has_key('FigureLinewidth'):           self.figureparameters['linewidth']  = kwargs['FigureLinewidth']
        if kwargs.has_key('FigureColorFull'):           self.figureparameters['colors'][0]  = kwargs['FigureColorFull']
        if kwargs.has_key('FigureColorEmpty'):          self.figureparameters['colors'][1]  = kwargs['FigureColorEmpty']
        if kwargs.has_key('FigureColorBackground'):     self.figureparameters['colors'][2]  = kwargs['FigureColorBackground']
        if kwargs.has_key('FigureColorBorder'):         self.figureparameters['colors'][3]  = kwargs['FigureColorBorder']
        if kwargs.has_key('FigureColorBorderNoGrowth'): self.figureparameters['colors'][4]  = kwargs['FigureColorBorderNoGrowth']



    def write_PNG(self,dataid,outfilename = None, growththreshold = None):
        if 0 <= dataid < len(self.__data):
            if outfilename is None:
                outfilename = self.__sheetnames[dataid].replace(' ','_') + '.png'
            else:
                if outfilename[-4:].upper() != '.PNG':
                    outfilename += '.png'
                
            cFull           = self.rgb(self.figureparameters['colors'][0])
            cEmpty          = self.rgb(self.figureparameters['colors'][1])
            cBorder         = self.rgb(self.figureparameters['colors'][2])
            cBack           = self.rgb(self.figureparameters['colors'][3])
            cBorderNoGrowth = self.rgb(self.figureparameters['colors'][4])

            CairoImage      = cairo.ImageSurface(cairo.FORMAT_ARGB32,self.__data[dataid].shape[1] * self.figureparameters['wellsize'],self.__data[dataid].shape[0] * self.figureparameters['wellsize'])
            context         = cairo.Context(CairoImage)

            context.rectangle(0,0,self.__data[dataid].shape[1] * self.figureparameters['wellsize'],self.__data[dataid].shape[0] * self.figureparameters['wellsize'])
            context.set_source_rgb(cBack[0],cBack[1],cBack[2])
            context.fill_preserve()
            context.new_path()

            context.set_line_width(self.figureparameters['linewidth'])
            context.translate(.5 * self.figureparameters['wellsize'],.5 * self.figureparameters['wellsize'])
            datamax   = np.amax(self.__data[dataid])
            datarange = np.amax(self.__data[dataid]) - np.amin(self.__data[dataid])
            if not growththreshold is None:
                threshold = (datamax - growththreshold)/datarange
            else:
                threshold = -1
            for x in range(int(self.__data[dataid].shape[1])):
                for y in range(int(self.__data[dataid].shape[0])):
                    context.new_path()
                    context.arc(0,0,self.figureparameters['wellradius'],0,2*math.pi)
                    r = (datamax - self.__data[dataid][y,x])/datarange
                    c = cFull * (1 - r) + cEmpty * r
                    context.set_source_rgb(c[0],c[1],c[2])
                    context.fill_preserve()
                    if r > threshold:
                        context.set_source_rgb(cBorderNoGrowth[0],cBorderNoGrowth[1],cBorderNoGrowth[2])
                    else:
                        context.set_source_rgb(cBorder[0],cBorder[1],cBorder[2])
                    context.stroke_preserve()
                    context.translate(0,self.figureparameters['wellsize'])
                context.translate(self.figureparameters['wellsize'],-self.__data[dataid].shape[0] * self.figureparameters['wellsize'])
            
            CairoImage.write_to_png(outfilename)

    def Export_All_PNGs(self):
        for i in range(self.count):
            self.write_PNG(i)

    def get_design(self,designid = 0, designname = None, dataID = None):
        if not designname is None:
            if designname in self.__designtitle:
                return self.__designdata[self.__designtitle.index(designname)]
        elif not dataID is None:
            return self.__designdata[self.__designassignment[dataID]]
        else:
            return self.__designdata[designid]
    
    
    def get_designassignment(self,dataid = None):
        if dataid is None:
            return self.__designassignment
        else:
            if 0 <= dataid < len(self.__designassignment):
                return self.__designassignment[dataid]
            else:
                raise KeyError
    
    
    def count_design(self):
        assert len(self.__designdata) == len(self.__designtitle)
        return len(self.__designdata)


    def add_design(self,xstart=6e6,xdilution = 4.,ystart=6.25,ydilution = 2,xdirection = 1,ydirection = -1,xsize = 8, ysize = 12, designtitle = 'generated'):
        if ydirection < 0:  ydirection = -1
        else:               ydirection =  1
        if xdirection < 0:  xdirection = -1
        else:               xdirection =  1
        if designtitle == 'generated':
            designtitle += '_X{}_dil{}_Y{}_dil{}'.format(xstart,xdilution,ystart,ydilution)
        design_x = np.array([[xstart * np.power(xdilution,-i) for j in np.arange(ysize)[::ydirection]] for i in np.arange(xsize)[::xdirection]],dtype=np.float)
        design_y = np.array([[ystart * np.power(ydilution,-j) for j in np.arange(ysize)[::ydirection]] for i in np.arange(xsize)[::xdirection]],dtype=np.float)
        
        self.__designdata.append((design_x,design_y))
        self.__designtitle.append(designtitle)

    def all_values(self):
        return np.concatenate([x.flatten() for x in self.__data])
    


    def interpolate_log_log(self,x1,x2,y1,y2,ythreshold):
        #print "{:.6e} {:.6e} {:.6e} {:.6e} {:.6e}".format(x1,x2,y1,y2,ythreshold)
        if x1 != x2:
            return x1 * np.exp( np.log(ythreshold/y1) * np.log(x2/x1) / np.log(y2/y1) )
        else:
            return x1


    def compute_growth_nogrowth_transition(self,dataid,threshold,design = 0, geom = True):
        r = list()
        platesize = np.shape(self.__data[dataid])
        allcont = measure.find_contours(self.__data[dataid],threshold)
        for cont in allcont:
            for pos in cont:
                    # logarithmic interpolation
                    i, j  = int(pos[0]//1), int(pos[1]//1)
                    di,dj = pos[0] - i, pos[1] - j

                    x = self.__designdata[design][0][i,j]
                    if j + 1 < platesize[1] and dj > 0:
                        x *= np.power(self.__designdata[design][0][i,j+1]/self.__designdata[design][0][i,j],dj)
                    
                    y = self.__designdata[design][1][i,j]
                    if i + 1 < platesize[0] and di > 0:
                        y *= np.power(self.__designdata[design][1][i+1,j]/self.__designdata[design][1][i,j],di)
                        
                    r.append(np.array([x,y]))

        if len(r) > 0:
            return np.vstack(r)
        else:
            return None


    def transitions(self,threshold):
        return [(self.__filenames[i],self.__sheetnames[i],self.compute_growth_nogrowth_transition(i,threshold,self.__designassignment[i])) for i in range(len(self.__data))]
    
    
    
    def write_data_to_file(self,dataID,filename = 'out',include_error_estimates = False):
        xlist = self.__designdata[self.__designassignment[dataID]][0]
        ylist = self.__designdata[self.__designassignment[dataID]][1]
        zlist = self.__data[dataID]
        if include_error_estimates:
            elist = self.get_noise_estimates(dataID)
        s = np.shape(zlist)
        
        fp = open(filename,'w')
        
        for i in range(s[0]):
            for j in range(s[1]):
                if include_error_estimates:
                    fp.write('{} {} {} {}\n'.format(xlist[i,j],ylist[i,j],zlist[i,j],elist[i,j]))
                else:
                    fp.write('{} {} {}\n'.format(xlist[i,j],ylist[i,j],zlist[i,j]))
            fp.write('\n')
        
        fp.close()

    
    def get_noise_estimates(self,dataID):
        # get rough estimate of noise as variance between neighboring wells on plate
        shape = np.shape(self.__data[dataID])
        ne = np.zeros(shape)

        #corners
        ne[0,0]                   = np.std(self.__data[dataID][:2,:2])
        ne[0,shape[1]-1]          = np.std(self.__data[dataID][:2:,-2:])
        ne[shape[0]-1,0]          = np.std(self.__data[dataID][-2:,:2])
        ne[shape[0]-1,shape[1]-1] = np.std(self.__data[dataID][-2:,-2:])

        # edges
        for i in range(1,shape[0]-1):
            ne[i,0]               = np.std(self.__data[dataID][i-1:i+1,:2])
            ne[i,shape[1]-1]      = np.std(self.__data[dataID][i-1:i+1,-2:])
        
        # edges
        for j in range(1,shape[1]-1):
            ne[0,j]               = np.std(self.__data[dataID][:2,j-1:j+1])
            ne[shape[0]-1,j]      = np.std(self.__data[dataID][-2:,j-1:j+1])
            
            # bulk
            for i in range(1,shape[0]-1):
                ne[i,j]           = np.std(self.__data[dataID][i-1:i+1,j-1:j+1])
        
        return ne


    def EstimateGrowthThreshold(self,dataID = None,historange = None, bins = None, logscale = True):
        # otsu's method
        # described in IEEE TRANSACTIONS ON SYSTEMS, MAN, AND CYBERNETICS (1979)
        # usually used to binarize photos into black/white, here we separate the growth/no-growth transition
        
        x = list()
        if dataID is None:
            # all data is used
            for i in range(len(self)):
                x += list(self.__data[i].flatten())
        elif isinstance(dataID,(list)):
            # pick specific IDs, provided as list
            for i in dataID:
                x += list(self.__data[i].flatten())
        elif isinstance(dataID,int):
            # estimate threshold only from a single plate
            x = list(self.__data[dataID])
        else:
            # something went wrong
            raise TypeError
        x = np.array(x)
        if logscale: x = np.log(x)

        if historange is None:
            if bins is None:
                count,binedges = np.histogram(x)
            else:
                count,binedges = np.histogram(x,bins=bins)
        else:
            if bins is None:
                count,binedges = np.histogram(x,range=historange)
            else:
                count,binedges = np.histogram(x,range=historange,bins=bins)
        bincenter = binedges[:-1] + .5 * np.diff(binedges)
        
        p   = count/float(sum(count))
        w   = np.array([np.sum(p[:k]) for k in range(bins)])
        m   = np.array([np.dot(p[:k],bincenter[:k]) for k in range(bins)])
        mT  = np.dot(p,bincenter)
        
        sB  = np.array([(mT * w[k] - m[k])**2/(w[k]*(1.-w[k])) if w[k]*(1.-w[k]) > 0 else 0 for k in range(bins)])
        idx = np.argmax(sB)
        
        if logscale:
            return np.exp(bincenter[idx])
        else:
            return bincenter[idx]
        
    
    
    def GaussianProcessRegression(self,dataID):
        #dummy
        return None
    
    
    
    
    
    
    
    
    

    def __iter__(self):
        for fn,title,platedata,designassignment in zip(self.__filenames,self.__sheetnames,self.__data,self.__designassignment):
            yield fn,title,self.rescale(platedata),designassignment
    
    
    def __int__(self):
        return len(self.__data)

    def __len__(self):
        return len(self.__data)
    
    def __getattr__(self,key):
        if key == "count":
            return len(self.__data)
        elif key == "count_design":
            return len(self.__designdata)
        elif key == "titles":
            return self.__sheetnames
        elif key == "filenames":
            return self.__filenames
        else:
            super(PlateReaderData,self).__getitem__(key)
    
    def __getitem__(self,key):
        return self.__data[key]
